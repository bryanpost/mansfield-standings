# engine.py — Mansfield League standings engine
#
# All qualification logic follows:
#   "DETERMINATION OF MANSFIELD EVENT PLAYOFF PARTICIPANTS – 2022-23"
#   (Current MCC League Guidelines.pdf, pages 5–6)
#
# Called by server.py on every GET /api/data request.
# Returns a dict that is serialised to JSON and consumed by the JS frontend
# (app.js), which is responsible only for rendering — no math lives there.

import csv as csv_mod
import pathlib

# openpyxl is only imported when an .xlsx file is actually opened
# (keeps the CSV-only path free of the dependency)


# ── Constants ─────────────────────────────────────────────────────────────────

GAMES_PER_HALF  = 9       # scheduled games per half per team
HALF_DONE_AVG   = 8       # avg W+L games threshold for "half is complete"
DRAW_RANK_X     = 73.0    # 'X' = not completed by deadline; ranked below 72"
DRAW_RANK_BLANK = 100.0   # blank = no score recorded; worst possible


# ── CSV parsing ───────────────────────────────────────────────────────────────

def parse_csv(results_path, config_path=None):
    """Read results.csv (and optionally config.csv) → (configs, teams).

    results.csv columns:
      shift, team, draw,
      h1_1 … h1_9,   ← one column per game slot in the first half
      h2_1 … h2_9    ← one column per game slot in the second half

      Each game cell contains W, L, B, or is blank (game not yet played).
      The draw column contains a distance in inches (e.g. 34 or 3.25),
      the letter X (not completed by deadline), or is blank (no score).

    config.csv columns:
      shift_id, label, h1_spots, h2_spots, wc_spots

    If config_path is None, the function looks for config.csv in the same
    directory as results.csv.
    """
    results_path = pathlib.Path(results_path)
    if config_path is None:
        config_path = results_path.parent / "config.csv"

    # Read config
    configs = []
    with open(config_path, newline='', encoding='utf-8') as f:
        for row in csv_mod.DictReader(f):
            configs.append({
                "shift_id": row["shift_id"].strip(),
                "label":    row["label"].strip(),
                "h1_spots": int(row["h1_spots"] or 0),
                "h2_spots": int(row["h2_spots"] or 0),
                "wc_spots": int(row["wc_spots"] or 0),
            })

    # Read per-game results
    teams = []
    with open(results_path, newline='', encoding='utf-8') as f:
        for row in csv_mod.DictReader(f):
            if not row.get("shift") or not row.get("team"):
                continue
            h1 = [_parse_cell(row.get(f"h1_{i}", "")) for i in range(1, 10)]
            h2 = [_parse_cell(row.get(f"h2_{i}", "")) for i in range(1, 10)]
            teams.append({
                "name":  row["team"].strip(),
                "shift": row["shift"].strip(),
                "h1":    [g for g in h1 if g],   # drop empty (unplayed) slots
                "h2":    [g for g in h2 if g],
                "draw":  _parse_draw(row.get("draw", "")),
            })

    return configs, teams


def _xl_sheet(label):
    """Sanitize a shift label into a valid Excel sheet name (removes ':')."""
    return label.replace(':', '-')


def parse_xlsx(results_path, config_path=None):
    """Read results.xlsx → (configs, teams).

    Each sheet corresponds to one shift.  Sheet names must match the 'label'
    column in config.csv (e.g. "Tuesday 4:25", "Wednesday Late").

    Columns per sheet (same as results.csv minus the 'shift' column):
      team, draw, h1_1 … h1_9, h2_1 … h2_9
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required: pip install openpyxl")

    results_path = pathlib.Path(results_path)
    if config_path is None:
        config_path = results_path.parent / "config.csv"

    # Read config
    configs = []
    with open(config_path, newline='', encoding='utf-8') as f:
        for row in csv_mod.DictReader(f):
            configs.append({
                "shift_id": row["shift_id"].strip(),
                "label":    row["label"].strip(),
                "h1_spots": int(row["h1_spots"] or 0),
                "h2_spots": int(row["h2_spots"] or 0),
                "wc_spots": int(row["wc_spots"] or 0),
            })

    # Sheet names sanitize ':' → '-' (Excel disallows ':' in sheet names)
    label_to_shift = {_xl_sheet(c["label"]): c["shift_id"] for c in configs}

    wb = openpyxl.load_workbook(results_path, read_only=True, data_only=True)

    teams = []
    for sheet_name in wb.sheetnames:
        shift_id = label_to_shift.get(sheet_name)
        if shift_id is None:
            continue  # ignore sheets not in config (e.g. a Notes sheet)

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # First row = column headers
        headers = [str(h).strip() if h is not None else '' for h in rows[0]]

        for row_vals in rows[1:]:
            row = dict(zip(headers, row_vals))
            team_name = str(row.get("team", "") or "").strip()
            if not team_name:
                continue
            h1 = [_parse_cell(row.get(f"h1_{i}", "")) for i in range(1, 10)]
            h2 = [_parse_cell(row.get(f"h2_{i}", "")) for i in range(1, 10)]
            teams.append({
                "name":  team_name,
                "shift": shift_id,
                "h1":    [g for g in h1 if g],
                "h2":    [g for g in h2 if g],
                "draw":  _parse_draw(row.get("draw", "")),
            })

    wb.close()
    return configs, teams


def csv_to_xlsx(csv_path, config_path, xlsx_path):
    """One-time migration: convert results.csv → results.xlsx.

    Creates one sheet per shift (sheet name = the human-readable shift label).
    Columns: team, draw, h1_1 … h1_9, h2_1 … h2_9.
    Column headers are styled for readability (bold, frozen first row).
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("openpyxl is required: pip install openpyxl")

    # Read config for the ordered list of shifts
    shift_labels = {}   # shift_id → label
    shift_order  = []
    with open(config_path, newline='', encoding='utf-8') as f:
        for row in csv_mod.DictReader(f):
            sid = row["shift_id"].strip()
            shift_labels[sid] = row["label"].strip()
            shift_order.append(sid)

    # Read CSV rows grouped by shift (preserving raw strings for each cell)
    shift_rows = {sid: [] for sid in shift_order}
    with open(csv_path, newline='', encoding='utf-8') as f:
        for row in csv_mod.DictReader(f):
            sid = row.get("shift", "").strip()
            if sid in shift_rows:
                shift_rows[sid].append(row)

    # Column structure
    game_cols = [f"h1_{i}" for i in range(1, 10)] + [f"h2_{i}" for i in range(1, 10)]
    headers   = ["team", "draw"] + game_cols

    # Style helpers
    hdr_font    = Font(bold=True)
    h1_fill     = PatternFill("solid", fgColor="DDEEFF")   # light blue  — H1
    h2_fill     = PatternFill("solid", fgColor="DDFDE0")   # light green — H2
    center      = Alignment(horizontal="center")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove the default empty sheet

    for sid in shift_order:
        label = shift_labels[sid]
        ws    = wb.create_sheet(title=_xl_sheet(label))

        # Header row
        ws.append(headers)
        for col_idx, h in enumerate(headers, start=1):
            cell          = ws.cell(row=1, column=col_idx)
            cell.font     = hdr_font
            cell.alignment = center
            if h.startswith("h1_"):
                cell.fill = h1_fill
            elif h.startswith("h2_"):
                cell.fill = h2_fill

        # Data rows
        for row in shift_rows.get(sid, []):
            ws.append([row.get(h, "") for h in headers])
            # Centre the game cells
            data_row = ws.max_row
            for col_idx in range(3, len(headers) + 1):
                ws.cell(row=data_row, column=col_idx).alignment = center

        # Column widths
        ws.column_dimensions["A"].width = 20   # team name
        ws.column_dimensions["B"].width = 8    # draw
        for col_letter in "CDEFGHIJKLMNOPQRSTUVW":
            ws.column_dimensions[col_letter].width = 4

        # Freeze header row
        ws.freeze_panes = "A2"

    wb.save(xlsx_path)


def _parse_cell(val):
    """Normalise one game-result cell → 'W', 'L', 'B', or '' (unplayed)."""
    v = str(val).strip().upper()
    if v in ('W', 'L', 'B'):
        return v
    return ''   # blank, underscore, anything else = not yet played


def _parse_draw(val):
    """Normalise a draw-shot value → float, 'X', or None."""
    s = str(val).strip().upper()
    if s in ('', 'NONE', '-', '—'):
        return None
    if s == 'X':
        return 'X'
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


# ── Record helpers ────────────────────────────────────────────────────────────

def hrec(games):
    """Wins and losses for one half's game list.  Byes do not count."""
    w = sum(1 for g in games if g == 'W')
    l = sum(1 for g in games if g == 'L')
    return {"w": w, "l": l}


def hrem(games):
    """Unplayed game slots remaining in one half.

    We assume GAMES_PER_HALF (9) total slots.  parse_seq only keeps played
    entries, so remaining = 9 − number of played entries.
    """
    played = sum(1 for g in games if g in ('W', 'L', 'B'))
    return max(0, GAMES_PER_HALF - played)


def orec(team):
    """Overall (H1 + H2 combined) win / loss record."""
    r1 = hrec(team["h1"])
    r2 = hrec(team["h2"])
    return {"w": r1["w"] + r2["w"], "l": r1["l"] + r2["l"]}


def pct(w, l):
    """Win percentage.  Returns 0.0 if no games played."""
    return w / (w + l) if (w + l) > 0 else 0.0


def games_behind(r1, r2):
    """How many games r1 is ahead of r2.

    Positive  → r1 leads.   Negative → r1 trails.
    Standard baseball formula: ((W1 − W2) + (L2 − L1)) / 2

    PDF example: 12-6 vs 11-6  →  GB = ((12-11)+(6-6))/2 = 0.5  →  "tied"
                 9-0  vs 8-1   →  GB = ((9-8)+(1-0))/2   = 1.0  →  NOT tied
    """
    return ((r1["w"] - r2["w"]) + (r2["l"] - r1["l"])) / 2.0


def ahead(r1, r2):
    """True if r1 is at least one FULL game ahead of r2 (GB ≥ 1.0).

    The rules treat teams within one half-game (GB < 1.0) as 'tied'.
    """
    return games_behind(r1, r2) >= 1.0


# ── Draw-shot helpers ─────────────────────────────────────────────────────────

def draw_rank(team):
    """Numeric sort key for draw-shot tiebreaker.  Lower = better.

    Ranking per rules (PDF p.7):
        numeric (0"–72") < X (missed deadline, ranked below 72") < blank (worst)
    """
    d = team["draw"]
    if d is None:
        return DRAW_RANK_BLANK
    if d == "X":
        return DRAW_RANK_X
    try:
        return float(d)
    except (ValueError, TypeError):
        return DRAW_RANK_BLANK


def draw_display(team):
    """Human-readable draw-shot string: '34"', '3¾"', 'X', or '—'."""
    d = team["draw"]
    if d is None:
        return "—"
    if d == "X":
        return "X"
    try:
        n = float(d)
    except (ValueError, TypeError):
        return "—"

    whole = int(n)
    frac  = round(n - whole, 3)
    frac_map = {0.375: "³⁄₈", 0.75: "¾", 0.5: "½", 0.25: "¼"}
    suffix = frac_map.get(frac, "")
    if suffix:
        return f'{whole}{suffix}"'
    if frac == 0.0:
        return f'{whole}"'
    return f'{n}"'


# ── Half-completion checks ────────────────────────────────────────────────────

def _avg_wl(teams, half_key):
    """Average number of W+L (non-bye) games played across a shift for one half."""
    if not teams:
        return 0.0
    total = sum(hrec(t[half_key])["w"] + hrec(t[half_key])["l"] for t in teams)
    return total / len(teams)


def is_h1_done(teams):
    """True when the average H1 games played (W+L only) across the shift is ≥ 8."""
    return _avg_wl(teams, "h1") >= HALF_DONE_AVG


def is_h2_done(teams):
    """True when the average H2 games played (W+L only) across the shift is ≥ 8."""
    return _avg_wl(teams, "h2") >= HALF_DONE_AVG


# ── Tiebreaker helpers ────────────────────────────────────────────────────────

def _filter_within_one_game(group, rec_fn):
    """Keep only teams within one full game of the leader's record.

    PDF (p.5): "teams within one half game of each other will be considered tied."
    A team is in the tied group if games_behind(leader, team) < 1.0.
    """
    if len(group) <= 1:
        return group
    leader_rec = max(
        (rec_fn(t) for t in group),
        key=lambda r: (pct(r["w"], r["l"]), r["w"]),
    )
    return [t for t in group if abs(games_behind(leader_rec, rec_fn(t))) < 1.0]


def _break_tie_half(tied_on_half):
    """Pick one winner from a group tied on the half record.

    Tiebreaker order — PDF p.6, "Tiebreaker Criteria for Half Winners and Runners-up":
      TB1  Overall season record, ≥ 1 full game  (filtered before this call,
             then applied again here via _filter_within_one_game on overall)
      TB2  H2H for the half            — NOT IMPLEMENTED
             (requires per-game opponent data not stored in standings.xlsx)
      TB3  H2H full season             — NOT IMPLEMENTED (same reason)
      TB4  Lowest draw-shot measurement
      TB5  Coin flip  →  approximated by alphabetical name order
    """
    # TB1: among teams tied on the half, further filter by overall record
    g = _filter_within_one_game(tied_on_half, orec)
    if len(g) == 1:
        return g[0]
    # TB4 + TB5
    return min(g, key=lambda t: (draw_rank(t), t["name"]))


def _break_tie_ewc(tied_group):
    """Pick one winner from a group tied on overall record for an Event WC spot.

    Tiebreaker order — PDF p.6, "Tiebreaker Criteria for Event Wild Cards":
      1.  Lowest draw-shot measurement
      2.  Coin flip  →  approximated by alphabetical name order
    """
    return min(tied_group, key=lambda t: (draw_rank(t), t["name"]))


# ── Core spot-allocation logic ────────────────────────────────────────────────

def determine_spots(pool, spots, rec_fn, rem_fn, half_done):
    """Determine which teams hold `spots` qualifying positions from `pool`.

    Runs spot-by-spot: each iteration finds the current leader, identifies
    all teams tied with them (within 1 game), picks the winner via tiebreakers,
    removes the winner from consideration, then moves to the next spot.

    Returns
    -------
    {
      "clinched":   set of team names that have *mathematically* secured a spot,
      "leading":    set of team names currently in a spot but not yet clinched,
      "clinch_how": dict  name → 'solo' | 'half' | 'overall'   (H1 only, for UI)
    }

    Clinch determination:
      Season in progress (half_done=False):
        Clinched if winner's WORST possible final record still beats the first
        non-qualifier's BEST possible final record by ≥ 1 full game.
      Half complete (half_done=True):
        Clinched if the winner leads the first non-qualifier by ≥ 1 full game
        on either the half record OR the overall record.
        The three 'clinch_how' labels explain *why*:
          'solo'    — no other team is within 1 game on the half (outright leader)
          'half'    — ≥1 game ahead on the half record vs first non-qualifier
          'overall' — tied within 1 game on half, but ≥1 game ahead overall
    """
    clinched  = set()
    leading   = set()
    clinch_how = {}

    if not pool or spots <= 0:
        return {"clinched": clinched, "leading": leading, "clinch_how": clinch_how}

    claimed = set()

    for _ in range(spots):
        available = [t for t in pool if t["name"] not in claimed]
        if not available:
            break

        # Sort by record (desc), then draw rank (asc) as secondary
        available.sort(key=lambda t: (-pct(rec_fn(t)["w"], rec_fn(t)["l"]), draw_rank(t)))

        # All teams tied with the leader within 1 full game on the *half* record
        tied_group = _filter_within_one_game(available, rec_fn)

        # Pick winner (tiebreakers applied inside _break_tie_half)
        winner = tied_group[0] if len(tied_group) == 1 else _break_tie_half(tied_group)
        claimed.add(winner["name"])

        # First team just outside the spots (the "cut line")
        remaining  = [t for t in available if t["name"] not in claimed]
        first_out  = remaining[0] if remaining else None

        if first_out is None:
            # Last team in the pool — automatically clinched
            clinched.add(winner["name"])
            clinch_how[winner["name"]] = "solo"
            continue

        if half_done:
            wr  = rec_fn(winner)
            fo  = rec_fn(first_out)
            ov_w = orec(winner)
            ov_f = orec(first_out)

            alone        = len(tied_group) == 1
            half_lead    = ahead(wr, fo)
            overall_lead = ahead(ov_w, ov_f)

            if alone or half_lead or overall_lead:
                clinched.add(winner["name"])
                if alone:
                    clinch_how[winner["name"]] = "solo"
                elif half_lead:
                    clinch_how[winner["name"]] = "half"
                else:
                    clinch_how[winner["name"]] = "overall"
            else:
                leading.add(winner["name"])

        else:
            # Worst case for winner vs best case for first_out
            wr    = rec_fn(winner)
            fo    = rec_fn(first_out)
            w_rem = rem_fn(winner)
            o_rem = rem_fn(first_out)

            worst_winner_half = {"w": wr["w"],        "l": wr["l"] + w_rem}
            best_out_half     = {"w": fo["w"] + o_rem, "l": fo["l"]}

            if ahead(worst_winner_half, best_out_half):
                # Clinched on half record alone
                clinched.add(winner["name"])
            else:
                # Not ahead on half in worst case — check TB1 (overall season record).
                # If winner's worst-case overall is ≥1 full game ahead of first_out's
                # best-case overall, winner wins the tiebreak regardless of half result.
                w_ov = orec(winner)
                f_ov = orec(first_out)
                worst_winner_ov = {"w": w_ov["w"], "l": w_ov["l"] + w_rem}
                best_out_ov     = {"w": f_ov["w"] + o_rem, "l": f_ov["l"]}
                if ahead(worst_winner_ov, best_out_ov):
                    clinched.add(winner["name"])
                else:
                    leading.add(winner["name"])

    return {"clinched": clinched, "leading": leading, "clinch_how": clinch_how}


# ── Per-shift status computation ──────────────────────────────────────────────

def compute_statuses(cfg, teams):
    """Compute playoff status for every team in one shift.

    Pool priority (PDF p.5 — order matters):
      1. H1 spots:  full team pool, based on 1st-half record
      2. H2 spots:  pool = teams NOT holding an H1 spot, based on 2nd-half record
      3. Shift WC:  pool = teams not in H1 or H2 spots, based on overall record
                    (only wed_late has wc_spots > 0)

    Status values returned per team:
      'clinched'   — mathematically secured a playoff spot
      'leadH1'     — leading an H1 spot; not yet clinched
      'leadH2'     — leading an H2 or Shift-WC spot; not yet clinched
      'contention' — still mathematically alive for remaining spots
      'elim'       — mathematically eliminated from all remaining paths
    """
    h1_done = is_h1_done(teams)
    h2_done = is_h2_done(teams)

    # ── H1 ────────────────────────────────────────────────────────────────────
    h1_res = determine_spots(
        pool      = teams,
        spots     = cfg["h1_spots"],
        rec_fn    = lambda t: hrec(t["h1"]),
        rem_fn    = lambda t: hrem(t["h1"]),
        half_done = h1_done,
    )
    h1_clinched = h1_res["clinched"]
    h1_leading  = h1_res["leading"]
    h1_how      = h1_res["clinch_how"]
    h1_in_spot  = h1_clinched | h1_leading

    # ── H2 (teams not already in H1 pool) ─────────────────────────────────────
    h2_pool = [t for t in teams if t["name"] not in h1_in_spot]
    h2_res = determine_spots(
        pool      = h2_pool,
        spots     = cfg["h2_spots"],
        rec_fn    = lambda t: hrec(t["h2"]),
        rem_fn    = lambda t: hrem(t["h2"]),
        half_done = h2_done,
    )
    h2_clinched = h2_res["clinched"]
    h2_leading  = h2_res["leading"]
    h2_in_spot  = h2_clinched | h2_leading

    # ── Shift WC (teams not in H1 or H2 pools) ────────────────────────────────
    wc_clinched = set()
    wc_leading  = set()
    if cfg["wc_spots"] > 0:
        wc_pool = [t for t in teams
                   if t["name"] not in h1_in_spot and t["name"] not in h2_in_spot]
        wc_res = determine_spots(
            pool      = wc_pool,
            spots     = cfg["wc_spots"],
            rec_fn    = orec,
            rem_fn    = lambda t: hrem(t["h1"]) + hrem(t["h2"]),
            half_done = h2_done,
        )
        wc_clinched = wc_res["clinched"]
        wc_leading  = wc_res["leading"]
    wc_in_spot = wc_clinched | wc_leading

    # ── Rank within each group for "W1" / "W2" labels ─────────────────────────
    team_by_name   = {t["name"]: t for t in teams}
    h2pool_by_name = {t["name"]: t for t in h2_pool}

    def _rank(names, key_fn, lookup):
        return {n: i for i, n in enumerate(sorted(names, key=lambda n: key_fn(lookup[n])))}

    h1_rank = _rank(
        h1_in_spot,
        lambda t: (-pct(hrec(t["h1"])["w"], hrec(t["h1"])["l"]), draw_rank(t)),
        team_by_name,
    )
    h2_rank = _rank(
        h2_in_spot,
        lambda t: (-pct(hrec(t["h2"])["w"], hrec(t["h2"])["l"]), draw_rank(t)),
        h2pool_by_name,
    )
    wc_rank = _rank(
        wc_in_spot,
        lambda t: (-pct(orec(t)["w"], orec(t)["l"]), draw_rank(t)),
        team_by_name,
    )

    # ── Build output list ──────────────────────────────────────────────────────
    result = []
    for t in teams:
        name = t["name"]
        r1   = hrec(t["h1"])
        r2   = hrec(t["h2"])
        ov   = orec(t)

        base = {
            "name":         name,
            "shift":        t["shift"],
            "h1":           {**r1, "pct": round(pct(r1["w"], r1["l"]), 3)},
            "h2":           {**r2, "pct": round(pct(r2["w"], r2["l"]), 3)},
            "overall":      {**ov, "pct": round(pct(ov["w"], ov["l"]), 3)},
            "draw_display": draw_display(t),
            "draw_rank":    draw_rank(t),   # numeric, for JS visual sorting
            "spot_label":   None,
            "h1_clinch_how": None,
            "status":       None,
        }

        if name in h1_in_spot:
            rank = h1_rank.get(name, 0)
            base["spot_label"]     = f"H1 W{rank + 1}"
            base["h1_clinch_how"]  = h1_how.get(name)
            base["status"]         = "clinched" if name in h1_clinched else "leadH1"

        elif name in h2_in_spot:
            rank = h2_rank.get(name, 0)
            base["spot_label"] = f"H2 W{rank + 1}"
            base["status"]     = "clinched" if name in h2_clinched else "leadH2"

        elif name in wc_in_spot:
            rank = wc_rank.get(name, 0)
            base["spot_label"] = f"SWC {rank + 1}"
            base["status"]     = "clinched" if name in wc_clinched else "leadH2"

        else:
            # Elimination check (H2 path only):
            # Eliminated if ≥ h2_spots other teams in the H2 pool each have their
            # WORST case beating this team's BEST case on H2 record.
            if name in h2pool_by_name:
                my_r2   = hrec(t["h2"])
                my_rem  = hrem(t["h2"])
                my_best = {"w": my_r2["w"] + my_rem, "l": my_r2["l"]}

                blockers = sum(
                    1 for other in h2_pool
                    if other["name"] != name
                    and ahead(
                        {"w": hrec(other["h2"])["w"],
                         "l": hrec(other["h2"])["l"] + hrem(other["h2"])},
                        my_best,
                    )
                )
                base["status"] = "elim" if blockers >= cfg["h2_spots"] else "contention"
            else:
                base["status"] = "contention"

        result.append(base)

    return result


# ── Collect all 28 playoff qualifiers ────────────────────────────────────────

def collect_qualified(configs, all_teams):
    """Collect all 28 playoff qualifiers (leaders + clinched) seeded by overall record.

    Qualification order per PDF p.5 (order of allocation matters — each step
    removes teams from subsequent pools):
      H1 winners → H1 runners-up → H2 winners → H2 runners-up →
      Shift Wild Cards → Event Wild Cards (4)

    The function runs the same determine_spots logic as compute_statuses,
    but across all shifts to build the global 28-team field.

    Returns a list of up to 28 team dicts, each augmented with:
      seed        — 1-based seed (overall record ranking)
      qual_how    — 'H1' | 'H2' | 'Shift WC' | 'Event WC'
      shift_label — human-readable shift name
      ov          — overall {w, l} record (used for seeding)
    """
    qualified  = []
    qual_names = set()

    for cfg in configs:
        shift_teams = [t for t in all_teams if t["shift"] == cfg["shift_id"]]
        if not shift_teams:
            continue

        h1_done = is_h1_done(shift_teams)
        h2_done = is_h2_done(shift_teams)

        # H1
        h1_res    = determine_spots(shift_teams, cfg["h1_spots"],
                                    lambda t: hrec(t["h1"]), lambda t: hrem(t["h1"]),
                                    h1_done)
        h1_in_spot = h1_res["clinched"] | h1_res["leading"]
        for name in h1_in_spot:
            if name not in qual_names:
                qual_names.add(name)
                team = next(t for t in shift_teams if t["name"] == name)
                qualified.append({**team, "ov": orec(team),
                                  "qual_how": "H1", "shift_label": cfg["label"]})

        # H2
        h2_pool = [t for t in shift_teams if t["name"] not in h1_in_spot]
        h2_res   = determine_spots(h2_pool, cfg["h2_spots"],
                                   lambda t: hrec(t["h2"]), lambda t: hrem(t["h2"]),
                                   h2_done)
        h2_in_spot = h2_res["clinched"] | h2_res["leading"]
        for name in h2_in_spot:
            if name not in qual_names:
                qual_names.add(name)
                team = next(t for t in shift_teams if t["name"] == name)
                qualified.append({**team, "ov": orec(team),
                                  "qual_how": "H2", "shift_label": cfg["label"]})

        # Shift WC
        if cfg["wc_spots"] > 0:
            wc_pool = [t for t in shift_teams
                       if t["name"] not in h1_in_spot and t["name"] not in h2_in_spot]
            wc_res  = determine_spots(wc_pool, cfg["wc_spots"],
                                      orec, lambda t: hrem(t["h1"]) + hrem(t["h2"]),
                                      h2_done)
            wc_in_spot = wc_res["clinched"] | wc_res["leading"]
            for name in wc_in_spot:
                if name not in qual_names:
                    qual_names.add(name)
                    team = next(t for t in shift_teams if t["name"] == name)
                    qualified.append({**team, "ov": orec(team),
                                      "qual_how": "Shift WC", "shift_label": cfg["label"]})

    # Event Wild Cards — best overall among ALL non-qualified teams
    # Tiebreaker: draw shot only (PDF p.6)
    ewc_pool = [
        {**t, "ov": orec(t)}
        for t in all_teams
        if t["name"] not in qual_names
    ]
    ewc_pool.sort(key=lambda t: (-pct(t["ov"]["w"], t["ov"]["l"]), draw_rank(t)))

    ewc_claimed = set()
    for _ in range(4):
        avail = [t for t in ewc_pool
                 if t["name"] not in ewc_claimed and t["name"] not in qual_names]
        if not avail:
            break
        leader_ov = avail[0]["ov"]
        tied      = [t for t in avail if abs(games_behind(leader_ov, t["ov"])) < 1.0]
        winner    = tied[0] if len(tied) == 1 else _break_tie_ewc(tied)
        ewc_claimed.add(winner["name"])
        qual_names.add(winner["name"])
        cfg_match = next((c for c in configs if c["shift_id"] == winner["shift"]), None)
        qualified.append({
            **winner,
            "qual_how":    "Event WC",
            "shift_label": cfg_match["label"] if cfg_match else winner["shift"],
        })

    # Seed 1–28 by overall record (PDF p.6 "Draw Principles": top seeds by EWC criteria)
    qualified.sort(key=lambda t: (-pct(t["ov"]["w"], t["ov"]["l"]), draw_rank(t)))
    for i, q in enumerate(qualified[:28]):
        q["seed"] = i + 1

    return qualified[:28]


# ── EWC pool for the standings-page display table ─────────────────────────────

def compute_ewc_pool(configs, all_teams, qualified):
    """Build the Event Wild Card display table for the standings page.

    Includes:
      - Teams that already hold an Event WC spot (shown with 'wc' badge)
      - All other non-qualified teams, tagged 'contention' or 'elim'

    A non-EWC team is 'elim' if the last current EWC holder's WORST case
    beats the contender's BEST case on overall record.
    """
    qual_names  = {t["name"] for t in qualified}
    ewc_winners = {t["name"] for t in qualified if t["qual_how"] == "Event WC"}

    # The cut line = the last (worst-seeded) Event WC team
    cut_team = next(
        (t for t in reversed(qualified) if t["qual_how"] == "Event WC"),
        None,
    )

    # Pool: non-qualified teams + teams holding an EWC spot (so they stay visible)
    pool = [
        {**t, "ov": orec(t)}
        for t in all_teams
        if t["name"] not in qual_names or t["name"] in ewc_winners
    ]
    pool.sort(key=lambda t: (-pct(t["ov"]["w"], t["ov"]["l"]), draw_rank(t)))

    result = []
    for t in pool:
        cfg = next((c for c in configs if c["shift_id"] == t["shift"]), None)
        entry = {
            "name":        t["name"],
            "shift":       t["shift"],
            "shift_label": cfg["label"] if cfg else t["shift"],
            "overall":     {**t["ov"], "pct": round(pct(t["ov"]["w"], t["ov"]["l"]), 3)},
            "draw_display": draw_display(t),
            "draw_rank":   draw_rank(t),
        }

        if t["name"] in ewc_winners:
            entry["ewc_status"] = "wc"
        elif cut_team is None:
            entry["ewc_status"] = "contention"
        else:
            cut_rem  = hrem(cut_team["h1"]) + hrem(cut_team["h2"])
            my_rem   = hrem(t["h1"])        + hrem(t["h2"])
            cut_worst = {"w": cut_team["ov"]["w"], "l": cut_team["ov"]["l"] + cut_rem}
            my_best   = {"w": t["ov"]["w"] + my_rem, "l": t["ov"]["l"]}
            entry["ewc_status"] = "elim" if ahead(cut_worst, my_best) else "contention"

        result.append(entry)

    return result


# ── Main entry point ──────────────────────────────────────────────────────────

def compute_all(results_path, config_path=None):
    """Read results.csv (and config.csv) and return the full JSON payload.

    Called on every GET /api/data request, so edits to results.csv are
    picked up automatically on the next browser refresh.
    """
    results_path = pathlib.Path(results_path)
    if results_path.suffix.lower() == '.xlsx':
        configs, all_teams = parse_xlsx(results_path, config_path)
    else:
        configs, all_teams = parse_csv(results_path, config_path)

    # Per-shift statuses
    shifts_out = []
    for cfg in configs:
        shift_teams = [t for t in all_teams if t["shift"] == cfg["shift_id"]]
        statuses    = compute_statuses(cfg, shift_teams)
        shifts_out.append({
            "shift_id": cfg["shift_id"],
            "label":    cfg["label"],
            "h1_spots": cfg["h1_spots"],
            "h2_spots": cfg["h2_spots"],
            "wc_spots": cfg["wc_spots"],
            "h1_done":  is_h1_done(shift_teams),
            "h2_done":  is_h2_done(shift_teams),
            "teams":    statuses,
        })

    # Global 28-team field
    qualified = collect_qualified(configs, all_teams)

    # EWC display table
    ewc_pool = compute_ewc_pool(configs, all_teams, qualified)

    # Serialise qualified list (strip internal game-list fields)
    qualified_out = [
        {
            "seed":        q["seed"],
            "name":        q["name"],
            "shift":       q["shift"],
            "shift_label": q["shift_label"],
            "overall":     {"w": q["ov"]["w"], "l": q["ov"]["l"],
                            "pct": round(pct(q["ov"]["w"], q["ov"]["l"]), 3)},
            "qual_how":    q["qual_how"],
        }
        for q in qualified
    ]

    return {
        "shifts":    shifts_out,
        "qualified": qualified_out,
        "ewc": {
            "winners_count":    sum(1 for t in ewc_pool if t["ewc_status"] == "wc"),
            "contention_count": sum(1 for t in ewc_pool if t["ewc_status"] == "contention"),
            "pool":             ewc_pool,
        },
    }
